import pathlib
import traceback
from typing import Any

import openpyxl

from invoice.core import file_io, info
from invoice.core import utilities as utils


class Excel_worker:
    def __init__(self, path: str | None = None, sheet: int | None = None):
        # Check if either both parameters are provided or neither is provided
        if (path is None) != (sheet is None):
            raise ValueError("Both 'path' and 'sheet' must be provided together or omitted")

        self.path = path
        self.sheet = sheet

        # private variables
        path_info = info.Path_info()
        self._instant_path = path_info.instance

    def write_cell(self, cell: str, value: Any, value_type: str = "string"):
        """
        Write excel file.

        Parameters:
        - cell (str): The cell address where the value will be written.
        - value (str, int, float): The value to be written to the cell.
        - value_type (str): The type of the value to be written. It can be one of the following:
            - 'string' (default)
            - 'float'
            - 'int'
            - 'scientific'
            - 'currency'

        Returns:
        None

        Raises:
        - Exception: If there is an error writing the excel file.

        Example:
        >>> write_cell('A1', 'Hello, World!')
        >>> write_cell('B2', 12345, 'number')"""
        try:
            wb_path = self._instant_path
            if pathlib.Path(wb_path).is_file() is False:
                raise ValueError("Excel file not instantiated!")
            if self.path is None or self.sheet is None:
                raise ValueError("Class must specify 'path' and 'sheet'!")

            wb = openpyxl.load_workbook(wb_path)
            sheet = wb.worksheets[self.sheet]
            cell = cell.upper()

            if value_type == "float":
                sheet[cell].number_format = "#,##0.00"
            elif value_type == "int":
                sheet[cell].number_format = "#,##0"
            elif value_type == "scientific":
                sheet[cell].number_format = "0.00E+00"
            elif value_type == "currency":
                sheet[cell].number_format = "$#,##0.00"
            sheet[cell] = value

            wb.save(wb_path)
        except Exception as e:
            print(f"Error writing excel: {e}")
            traceback.print_exc()

    def read_cell(self, cell: str):
        """read excel file"""

        try:
            wb_path = self._instant_path
            if pathlib.Path(wb_path).is_file() is False:
                raise ValueError("Excel file not instantiated!")
            if self.path is None or self.sheet is None:
                raise ValueError("Class must specify 'path' and 'sheet'!")

            wb = openpyxl.load_workbook(wb_path)
            sheet = wb.worksheets[self.sheet]
            value = sheet[cell].value
            return value
        except Exception as e:
            print(f"Error reading excel: {e}")
            traceback.print_exc()
            return None

    def get_last_data_row(self, column: str, start_row: int, row_range: int):
        """get last data row"""
        try:
            wb_path = self._instant_path
            if pathlib.Path(wb_path).is_file() is False:
                raise ValueError("Excel file not instantiated!")
            if self.path is None or self.sheet is None:
                raise ValueError("Class must specify 'path' and 'sheet'!")
            
            wb = openpyxl.load_workbook(wb_path)
            sheet = wb.worksheets[self.sheet]
            non_empty_rows = []
            for i in range(row_range):
                cell = utils.concat_pos(column, start_row + i)
                value = sheet[cell].value
                if value is not None:
                    non_empty_rows.append(start_row + i)

            try:
                last_row = non_empty_rows[-1]
            except IndexError:
                last_row = start_row - 1

            if len(non_empty_rows) >= row_range:
                print(f"non_empty_rows: {non_empty_rows}")
                raise ValueError(
                    f"Maximum row range exceeded! row_range: {row_range}, at row: {last_row}"
                )
            return last_row

        except Exception as e:
            print(f"Error getting last data row: {e}")
            traceback.print_exc()

    def remove_row(self, row_index: int, start_row: int, row_range: int):
        """
        Remove a row from the Excel file.

        Parameters:
        - row_index (int): The index of the row to be removed. If row_index is positive, it represents the absolute index of the row. If row_index is negative, it represents the relative index of the row with respect to the last data row.
        - start_row (int): The index of the first row in the range where the row can be removed.
        - row_range (int): The number of rows in the range where the row can be removed.

        Returns:
        None

        Raises:
        - ValueError: If the Excel file is not instantiated.
        - ValueError: If there is no data to remove at the specified row.
        - ValueError: If the row index is out of range.

        Example:
        >>> remove_row(2, 1, 10)
        >>> remove_row(-1, 1, 10) # remove last row"""
        try:
            wb_path = self._instant_path
            if pathlib.Path(wb_path).is_file() is False:
                raise ValueError("Excel file not instantiated!")
            if self.path is None or self.sheet is None:
                raise ValueError("Class must specify 'path' and 'sheet'!")

            wb = openpyxl.load_workbook(wb_path)
            sheet = wb.worksheets[self.sheet]

            if row_index >= 0:
                row_to_remove = start_row + row_index
            else:
                row_to_remove = self.get_last_data_row("B", start_row, row_range + 1)
                if row_to_remove is None:
                    raise ValueError(f"No data to remove at row {row_to_remove}!")
                row_to_remove += row_index + 1

            # ensure row_to_remove is within range
            if row_to_remove < start_row or row_to_remove > start_row + row_range:
                raise ValueError("Row index out of range!")
            for cell in sheet[row_to_remove]:
                print(cell.coordinate, cell.value)
                cell.value = None  # remove cell value

            wb.save(wb_path)
        except Exception as e:
            print(f"Error removing row: {e}")
            traceback.print_exc()

    def instantiate(self):
        """instantiate excel file"""
        try:
            if self.path is None or self.sheet is None:
                raise ValueError("Class must specify 'path' and 'sheet'!")
            
            wb = openpyxl.load_workbook(self.path)
            wb.save(self._instant_path)
            return self._instant_path
        except Exception as e:
            print(f"Error instantiating excel: {e}")
            traceback.print_exc()

    def is_instantiated(self):
        """check if excel file is instantiated"""
        if not pathlib.Path(self._instant_path).is_file():
            return False
        return True

    def clean_up(self):
        """clean up instant excel file"""
        try:
            file_io.delete_file(self._instant_path)
        except Exception as e:
            print(f"Error cleaning up: {e}")
            traceback.print_exc()
